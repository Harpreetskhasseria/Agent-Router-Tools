# app_combined2.py

import streamlit as st
import pandas as pd
from io import BytesIO
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from pathlib import Path
import asyncio
from urllib.parse import urlparse
import os

from phase1_web_pipeline import app as langgraph_app
from tools.scraper_tool import scraper_tool
from tools.cleaner_tool import cleaner_tool
from tools.html_extractor_tool import html_extractor_tool
from agents.summarizer_agent import SummarizerAgent

st.set_page_config(page_title="Combined Regulatory Analyzer", layout="wide")
st.title("🔄 Combined Regulatory Horizon Analyzer")

# Step 1: User enters URL
url_input = st.text_input("Enter Regulatory Website URL", placeholder="https://www.example.com/news")
run_button = st.button("▶️ Run Phase 1 (Extract + Exclude)")

output_excel_path = None

if run_button and url_input:
    with st.spinner("Running Phase 1 pipeline..."):
        try:
            result = asyncio.run(langgraph_app.ainvoke({
                "url": url_input,
                "scraper_input": {"url": url_input},
                "route": "web"
            }))

            output_excel_path = result.get("final_output", {}).get("output_file", "")
            st.success("✅ Phase 1 complete. Proceeding to summarization phase.")
        except Exception as e:
            st.error(f"❌ Failed to process URL: {str(e)}")

# Step 2: Display output table and run summarization
def extract_hyperlinks_from_excel_column(path, column_letter="G"):
    wb = load_workbook(path, data_only=False)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    links = []
    for cell in ws[column_letter]:
        val = cell.value
        if isinstance(val, str) and val.startswith('=HYPERLINK('):
            try:
                url = val.split('"')[1]
                if not url.startswith(('http://', 'https://')):
                    url = 'https://' + url
                links.append(url)
            except IndexError:
                links.append("")
        else:
            links.append("")
    return links

if output_excel_path:
    try:
        real_links = extract_hyperlinks_from_excel_column(output_excel_path, column_letter="G")
        df = pd.read_excel(output_excel_path, engine="openpyxl")
        df["Link"] = real_links[:len(df)] + [""] * (len(df) - len(real_links))

        if "action" not in df.columns:
            df["action"] = "no action"
        else:
            df["action"] = df["action"].astype(str).fillna("no action")

        action_options = ["no action", "summarize", "custom:<your prompt>"]
        df.loc[~df["action"].isin(action_options) & ~df["action"].str.startswith("custom:"), "action"] = "no action"

        if "phase2_output" not in df.columns:
            df["phase2_output"] = ""

        st.session_state["edited_df"] = df

    except Exception as e:
        st.error(f"⚠️ Could not load Phase 1 Excel: {str(e)}")

# Step 3: Show editor and summarization button
if "edited_df" in st.session_state:
    st.subheader("📋 Regulatory Updates - Select Actions and Run Summarization")

    st.session_state["edited_df"] = st.data_editor(
        st.session_state["edited_df"],
        column_config={
            "action": st.column_config.SelectboxColumn(
                "Action",
                options=["no action", "summarize", "custom:<your prompt>"],
                required=False
            ),
            "Link": st.column_config.LinkColumn(
                "Link",
                display_text="Click here"
            )
        },
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True
    )

    if st.button("▶️ Run Summarization for Selected Rows"):
        summarizer = SummarizerAgent()
        results = []
        df = st.session_state["edited_df"]
        new_outputs = []

        for idx, row in df.iterrows():
            action = str(row.get("action", "")).strip().lower()
            url = str(row.get("Link", "")).strip()
            existing = str(row.get("phase2_output", "")).strip().lower()

            if action.startswith("custom:") or action == "summarize":
                if existing not in ["", "skipped", "⏭️ skipped"]:
                    new_outputs.append(existing)
                    continue

                try:
                    scraped = scraper_tool.run(url=url)
                    cleaned = cleaner_tool.run(url=url, scraped_html=scraped["scraped_html"])
                    extracted = html_extractor_tool.run(url=url, cleaned_file=cleaned["cleaned_file"])

                    if action == "summarize":
                        summary = summarizer.run({
                            "extracted_text": extracted["extracted_text"],
                            "url": url
                        })
                    else:
                        prompt = action[7:].strip()
                        summarizer.goal = prompt
                        summary = summarizer.run({
                            "extracted_text": extracted["extracted_text"],
                            "url": url
                        })

                    new_outputs.append(summary.get("summary", "✅ Success but empty"))
                except Exception as e:
                    new_outputs.append(f"❌ Error: {str(e)}")
            else:
                new_outputs.append("⏭️ Skipped")

        st.session_state["edited_df"]["phase2_output"] = new_outputs
        st.rerun()

        # Save to disk
        output_path = Path("regulatory_outputs/site_outputs/phase2_summary_output.xlsx")
        st.session_state["edited_df"].to_excel(output_path, index=False)
        st.success("✅ Summarization complete. Table updated.")

# Final Download
if "edited_df" in st.session_state:
    output_buffer = BytesIO()
    st.session_state["edited_df"].to_excel(output_buffer, index=False, engine="openpyxl")
    output_buffer.seek(0)

    st.download_button(
        "⬇️ Download Results as Excel",
        output_buffer,
        file_name="phase2_summary_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
