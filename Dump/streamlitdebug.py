import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="🔗 Excel Link Debugger", layout="wide")
st.title("🔍 Extract Clickable Hyperlinks from Excel Formulas")

# --- Helper: Extract HYPERLINK() URLs from column G using openpyxl ---
def extract_hyperlinks_from_excel_column(path, column_letter="G"):
    wb = load_workbook(path, data_only=False)  # Get formulas, not just values
    first_sheet = wb.sheetnames[0]             # Automatically use the first sheet
    ws = wb[first_sheet]
    links = []
    for cell in ws[column_letter]:
        val = cell.value
        if isinstance(val, str) and val.startswith('=HYPERLINK('):
            try:
                url = val.split('"')[1]
                if not url.startswith(('http://', 'https://')):
                    url = 'https://' + url
                links.append(url)
            except IndexError:
                links.append("")
        else:
            links.append("")
    return links

# --- Step 1: Upload or specify Excel path ---
uploaded_file = st.file_uploader("📁 Upload Excel with hyperlinks (XLSX only)", type=["xlsx"])
manual_path = st.text_input("Or enter full Excel file path")

# --- Step 2: Load Excel if provided ---
if uploaded_file or manual_path:
    try:
        file_path = uploaded_file if uploaded_file else manual_path
        df = pd.read_excel(file_path, engine="openpyxl")
        file_label = uploaded_file.name if uploaded_file else Path(manual_path).name
    except Exception as e:
        st.error(f"❌ Failed to read Excel file: {e}")
        st.stop()

    st.success(f"✅ Loaded: {file_label}")

    # --- Step 3: Extract hyperlinks from column G in the actual Excel (openpyxl) ---
    real_links = extract_hyperlinks_from_excel_column(file_path, column_letter="G")

    # Match the number of links to DataFrame length
    real_links = real_links[:len(df)]  # Trim if too long
    real_links += [""] * (len(df) - len(real_links))  # Pad if too short
    df["Link"] = real_links


    # --- Step 4: DEBUG Output ---
    st.write("🔍 DEBUG: Extracted links")
    st.write(df["Link"].tolist())

    # --- Step 5: Handle 'action' column dropdown options ---
    action_options = ["no action", "summarize", "custom:<your prompt>"]
    if "action" not in df.columns:
        df["action"] = "no action"
    else:
        df["action"] = df["action"].astype(str).fillna("no action")
        df.loc[
            ~df["action"].isin(action_options)
            & ~df["action"].str.startswith("custom:"),
            "action"
        ] = "no action"

    # --- Step 6: Render Data Editor ---
    edited_df = st.data_editor(
        df,
        column_config={
            "action": st.column_config.SelectboxColumn("Action", options=action_options),
            "Link": st.column_config.LinkColumn("Link", display_text=None)
        },
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True
    )

    # --- Step 7: Allow download of modified file ---
    buffer = BytesIO()
    edited_df.to_excel(buffer, index=False, engine="openpyxl")
    st.download_button(
        "⬇️ Download Modified Excel",
        data=buffer.getvalue(),
        file_name=f"modified_{file_label}",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("📌 Please upload an Excel file or enter its full path above to continue.")
