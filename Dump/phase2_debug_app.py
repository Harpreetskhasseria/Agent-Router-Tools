import streamlit as st
import pandas as pd
import os
import asyncio
from urllib.parse import urlparse
from pathlib import Path
from openpyxl import load_workbook
from io import BytesIO

from agents.router_agent import RouterAgent
from phase1_web_pipeline import app as langgraph_app
from tools.scraper_tool import scraper_tool
from tools.cleaner_tool import cleaner_tool
from tools.html_extractor_tool import html_extractor_tool
from agents.summarizer_agent import SummarizerAgent

st.set_page_config(page_title="Regulatory Horizon Scanner", layout="wide")
st.title("📡 Regulatory Horizon Intelligence Platform")

# 🔧 Extract hyperlinks from Excel formulas (column G)
def extract_hyperlinks_from_excel_column(path, column_letter="G"):
    wb = load_workbook(path, data_only=False)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    links = []
    for cell in ws[column_letter]:
        val = cell.value
        if isinstance(val, str) and val.startswith('=HYPERLINK('):
            try:
                url = val.split('"')[1]
                if not url.startswith(('http://', 'https://')):
                    url = 'https://' + url
                links.append(url)
            except IndexError:
                links.append("")
        else:
            links.append("")
    return links

# PHASE 1: URL or RSS Input
url = st.text_input("Enter a regulatory website URL or RSS:")

if st.button("Run Phase 1"):
    if not url.strip():
        st.warning("⚠️ Please enter a valid URL.")
    else:
        with st.spinner("🔎 Routing the input..."):
            router = RouterAgent()
            route_info = router.run({"url": url})
            route = route_info.get("route", "web")

        if route != "web":
            st.warning("📡 RSS Flow is under construction. Please input a URL.")
        else:
            st.success("📬 Route selected: WEB (scraper pipeline)")
            with st.spinner("⚙️ Running Phase 1 pipeline..."):
                result = asyncio.run(langgraph_app.ainvoke({
                    "url": url,
                    "scraper_input": {"url": url},
                    "route": "web"
                }))

            parsed_domain = urlparse(url).netloc.replace(".", "_")
            search_folder = Path("regulatory_outputs/site_outputs")
            matched_files = sorted(
                search_folder.glob(f"{parsed_domain}_llm_exclusion_checked_*.xlsx"),
                key=os.path.getmtime,
                reverse=True
            )

            if matched_files:
                output_path = matched_files[0]
                real_links = extract_hyperlinks_from_excel_column(output_path, column_letter="G")
                df = pd.read_excel(output_path, engine="openpyxl")
                df["Link"] = real_links[:len(df)] + [""] * (len(df) - len(real_links))

                if "action" not in df.columns:
                    df["action"] = "no action"
                else:
                    df["action"] = df["action"].astype(str).fillna("no action")

                action_options = ["no action", "summarize", "custom:<your prompt>"]
                df.loc[~df["action"].isin(action_options) & ~df["action"].str.startswith("custom:"), "action"] = "no action"

                if "phase2_output" not in df.columns:
                    df["phase2_output"] = ""

                st.session_state["edited_df"] = df
                st.session_state["exclusion_file"] = str(output_path)
                st.success("✅ Phase 1 completed. You may now run summarization.")
            else:
                st.error("❌ Could not locate exclusion output file.")

# ✅ Show Phase 2 UI if DF loaded
if "edited_df" in st.session_state:
    st.subheader("📋 Regulatory Updates - Select Actions and Run Summarization")

    st.session_state["edited_df"] = st.data_editor(
        st.session_state["edited_df"],
        column_config={
            "action": st.column_config.SelectboxColumn(
                "Action",
                options=["no action", "summarize", "custom:<your prompt>"],
                required=False
            ),
            "Link": st.column_config.LinkColumn(
                "Link",
                display_text="Click here"
            )
        },
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True
    )

    if st.button("▶️ Run Summarization for Selected Rows"):
        summarizer = SummarizerAgent()
        df = st.session_state["edited_df"]
        new_outputs = []

        for idx, row in df.iterrows():
            action = str(row.get("action", "")).strip().lower()
            url = str(row.get("Link", "")).strip()
            existing = str(row.get("phase2_output", "")).strip().lower()

            if action.startswith("custom:") or action == "summarize":
                if existing not in ["", "skipped", "⏭️ skipped"]:
                    new_outputs.append(existing)
                    continue

                try:
                    scraped = scraper_tool.run(url=url)
                    cleaned = cleaner_tool.run(url=url, scraped_html=scraped["scraped_html"])
                    extracted = html_extractor_tool.run(url=url, cleaned_file=cleaned["cleaned_file"])

                    if action == "summarize":
                        summary = summarizer.run({
                            "extracted_text": extracted["extracted_text"],
                            "url": url
                        })
                    else:
                        prompt = action[7:].strip()
                        summarizer.goal = prompt
                        summary = summarizer.run({
                            "extracted_text": extracted["extracted_text"],
                            "url": url
                        })

                    new_outputs.append(summary.get("summary", "✅ Success but empty"))
                except Exception as e:
                    new_outputs.append(f"❌ Error: {str(e)}")
            else:
                new_outputs.append("⏭️ Skipped")

        st.session_state["edited_df"]["phase2_output"] = new_outputs
        output_path = Path("regulatory_outputs/site_outputs/phase2_summary_output.xlsx")
        st.session_state["edited_df"].to_excel(output_path, index=False)
        st.success("✅ Summarization complete. Table updated.")
        st.rerun()

# ✅ Download button
if "edited_df" in st.session_state:
    output_buffer = BytesIO()
    st.session_state["edited_df"].to_excel(output_buffer, index=False, engine="openpyxl")
    output_buffer.seek(0)

    st.download_button(
        "⬇️ Download Results as Excel",
        output_buffer,
        file_name="phase2_summary_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
