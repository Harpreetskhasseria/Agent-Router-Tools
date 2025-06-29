import streamlit as st
import pandas as pd
import os
import asyncio
from urllib.parse import urlparse
from pathlib import Path
from openpyxl import load_workbook
from agents.router_agent import RouterAgent
from phase1_web_pipeline import app as langgraph_app

# Phase 2 Tools and Agent
from tools.scraper_tool import scraper_tool
from tools.cleaner_tool import cleaner_tool
from tools.html_extractor_tool import html_extractor_tool
from agents.summarizer_agent import SummarizerAgent

st.set_page_config(page_title="Regulatory Horizon Scanner", layout="wide")
st.title("📡 Regulatory Horizon Intelligence Platform")

# 🔧 Extract hyperlinks from Excel formulas
def extract_hyperlinks_from_excel_column(path, column_letter="G"):
    wb = load_workbook(path, data_only=False)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    links = []
    for cell in ws[column_letter]:
        val = cell.value
        if isinstance(val, str) and val.startswith('=HYPERLINK('):
            try:
                url = val.split('"')[1]
                if not url.startswith(('http://', 'https://')):
                    url = 'https://' + url
                links.append(url)
            except IndexError:
                links.append("")
        else:
            links.append("")
    return links

# Phase 1: Input URL and run pipeline
url = st.text_input("Enter a regulatory website URL:")

if st.button("Run Phase 1"):
    if not url.strip():
        st.warning("⚠️ Please enter a valid URL.")
    else:
        with st.spinner("🔎 Routing the URL..."):
            router = RouterAgent()
            route_info = router.run({"url": url})
            route = route_info.get("route", "web")

        if route != "web":
            st.error("❌ RSS route is not yet supported. Please enter a website URL.")
        else:
            st.success("📬 Route selected: WEB (scraper pipeline)")
            with st.spinner("⚙️ Running Phase 1 pipeline..."):
                result = asyncio.run(langgraph_app.ainvoke({
                    "url": url,
                    "scraper_input": {"url": url},
                    "route": "web"
                }))

            parsed_domain = urlparse(url).netloc.replace(".", "_")
            search_folder = Path("regulatory_outputs/site_outputs")
            matched_files = sorted(
                search_folder.glob(f"{parsed_domain}_llm_exclusion_checked_*.xlsx"),
                key=os.path.getmtime,
                reverse=True
            )

            if matched_files:
                output_path = matched_files[0]
                df = pd.read_excel(output_path, engine="openpyxl")
                real_links = extract_hyperlinks_from_excel_column(output_path, column_letter="G")
                real_links = real_links[:len(df)] + [""] * (len(df) - len(real_links))
                df["Link"] = real_links

                df.rename(columns={
                    "date": "Date",
                    "topic": "Topic",
                    "additional_context": "Context",
                    "regulator": "Regulator"
                }, inplace=True)

                action_options = ["no action", "summarize", "custom:<your prompt>"]
                if "action" not in df.columns or df["action"].isnull().all():
                    df["action"] = "no action"
                else:
                    df["action"] = df["action"].astype(str).fillna("no action")
                    df.loc[~df["action"].isin(action_options) & ~df["action"].str.startswith("custom:"), "action"] = "no action"

                st.success("✅ Phase 1 completed. Review results and select actions:")
                st.markdown("""
                    <style>
                    .element-container:has(.stDataEditor) {
                        overflow-x: auto;
                    }
                    </style>
                """, unsafe_allow_html=True)

                with st.container():
                    edited_df = st.data_editor(
                        df,
                        column_config={
                            "action": st.column_config.SelectboxColumn(
                                "Action",
                                options=action_options,
                                required=False
                            ),
                            "Link": st.column_config.LinkColumn(
                                "Link",
                                display_text="Click here"
                            )
                        },
                        use_container_width=True,
                        num_rows="dynamic",
                        hide_index=True
                    )

                st.session_state["phase1_df"] = edited_df
                st.session_state["phase1_file"] = str(output_path)

                if st.button("💾 Save Actions for Phase 2"):
                    edited_df.to_excel(output_path, index=False)
                    st.success(f"✅ Actions saved to {output_path.name}")

                with open(output_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Phase 1 Excel",
                        f.read(),
                        file_name=output_path.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.error("❌ Could not find or open Phase 1 output file.")

# 🧠 Phase 2 Summarization
if "phase1_df" in st.session_state:
    st.subheader("🧠 Phase 2: Summarization")
    if st.button("▶️ Run Phase 2 Summarization"):
        with st.spinner("🔄 Running summarization for selected rows..."):
            phase1_df = st.session_state["phase1_df"].copy()
            summarizer = SummarizerAgent()
            summaries = []

            for _, row in phase1_df.iterrows():
                action = row.get("action", "").strip().lower()
                url = row.get("Link", "").strip()

                if action == "summarize" and url:
                    try:
                        scraped = scraper_tool.run(url=url)
                        cleaned = cleaner_tool.run(url=url, scraped_html=scraped["scraped_html"])
                        extracted = html_extractor_tool.run(url=url, cleaned_file=cleaned["cleaned_file"])
                        summary = summarizer.run({
                            "extracted_text": extracted["extracted_text"],
                            "url": url
                        })
                        summaries.append(summary.get("summary", "✅ Success but empty"))
                    except Exception as e:
                        summaries.append(f"❌ Error: {str(e)}")
                else:
                    summaries.append("⏭️ Skipped")

            phase1_df["phase2_output"] = summaries

            updated_output_path = Path(st.session_state["phase1_file"]).with_name(
                Path(st.session_state["phase1_file"]).stem + "_phase2_output.xlsx"
            )
            phase1_df.to_excel(updated_output_path, index=False)
            st.session_state["phase2_file"] = str(updated_output_path)

            st.success("✅ Phase 2 summarization complete.")
            st.dataframe(phase1_df, use_container_width=True)

            with open(updated_output_path, "rb") as f:
                st.download_button(
                    "⬇️ Download Phase 2 Excel",
                    f.read(),
                    file_name=updated_output_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
